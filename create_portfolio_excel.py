#!/usr/bin/env python3
"""
Script to create Portfolio Excel file with company data
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def create_portfolio_excel():
    # Create a new workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Portfolio Overview"
    
    # Define headers
    headers = [
        "Company",
        "Industry", 
        "Acquisition Year",
        "Revenue 2024 in kEUR",
        "FTE",
        "Headquarter",
        "Countries with branches",
        "Seller",
        "Website"
    ]
    
    # Portfolio data
    portfolio_data = [
        ["WPS", "Software", 2023, 27000, 115, "Netherlands", 5, "Non-core", "https://www.wpsparking.com"],
        ["MOBOTIX", "Software", 2025, 55000, 269, "Germany", 6, "Non-core", "https://www.mobotix.com"],
        ["AFT", "Industrials", 2007, 120000, 350, "Germany", 8, "Kardex AG", "https://www.aft-group.de"],
        ["COI", "Software", 2004, 5000, 55, "Germany", 1, "Schaeffler", "https://www.coi.de"],
        ["DEMM", "Industrials", 2018, 15000, 80, "Italy", 1, "Insolvency", "https://www.demm.it"],
        ["Enorise", "Automotive", 2024, 105000, 360, "USA", 5, "Non-core", "https://www.enorise.com"],
        ["Hannemann", "Industrials", 1999, 10000, 40, "Germany", 1, "Pruss family", "https://www.hannemann.de"],
        ["HermannKoch", "Packaging", 2015, 25000, 150, "Germany", 1, "Family", "https://www.hkoch.de"],
        ["Horn", "Industrials", 2000, 70000, 300, "Germany", 5, "n.a.", "https://www.hornglass.com"],
        ["ISOLITE", "Industrials", 2020, 70000, 500, "Germany", 7, "Mitsubishi", "https://www.isolite.de"],
        ["PAL", "Automotive", 2018, 45000, 550, "Czech Rep.", 1, "Pricol India", "https://www.pal.cz"],
        ["Pruss", "Industrials", 1997, 25000, 90, "Germany", 1, "Pruss family", "https://www.pruss.de"],
        ["Qcision", "Industrials", 2023, 40000, 140, "Switzerland", 1, "Feintool", "https://www.qcision.com"],
        ["Qualiform", "Packaging", 2015, 15000, 80, "France", 1, "Pochet", "https://www.qualiform.fr"],
        ["Rebhan", "Packaging", 2012, 30000, 220, "Germany", 3, "BLB Capital", "https://www.rebhan.de"],
        ["SCV", "Automotive", 2012, 10000, 30, "Italy", 1, "Amplifon", "https://www.scvspecial.com"],
        ["Shieldtech", "Industrials", 2024, 175000, 650, "Switzerland", 2, "ElringKlinger", "https://www.shieldtecgroup.com"],
        ["Trassl", "Packaging", 2020, 5000, 50, "Germany", 1, "Owner family", "https://www.trassl-polymer.de"],
        ["UST", "Goods & Services", 2021, 40000, 200, "Germany", 1, "Zwilling", "https://www.ust-germany.com"],
        ["VielfaltMenü", "Goods & Services", 2018, 25000, 160, "Germany", 1, "Apetito", "https://www.vielfaltmenue.com"],
        ["Wegener+Stapel", "Industrials", 2019, 15000, 50, "Germany", 1, "Owner family", "https://www.wegenerwelding.de"]
    ]
    
    # Add headers to worksheet
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E3A4A", end_color="2E3A4A", fill_type="solid")  # Certina dark blue
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data to worksheet
    for row_idx, row_data in enumerate(portfolio_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Format revenue column with thousand separators
            if col_idx == 4:  # Revenue column
                cell.number_format = '#,##0'
            
            # Format website column as hyperlinks
            if col_idx == 9:  # Website column
                cell.hyperlink = value
                cell.font = Font(color="0000FF", underline="single")
            
            # Alternate row colors for better readability
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add summary statistics at the bottom
    summary_row = len(portfolio_data) + 3
    
    # Summary headers
    ws.cell(row=summary_row, column=1, value="Summary Statistics").font = Font(bold=True, size=12)
    
    # Total companies
    ws.cell(row=summary_row + 1, column=1, value="Total Companies:")
    ws.cell(row=summary_row + 1, column=2, value=len(portfolio_data))
    
    # Total revenue
    ws.cell(row=summary_row + 2, column=1, value="Total Revenue (kEUR):")
    total_revenue = sum(row[3] for row in portfolio_data)
    ws.cell(row=summary_row + 2, column=2, value=total_revenue)
    ws.cell(row=summary_row + 2, column=2).number_format = '#,##0'
    
    # Total FTE
    ws.cell(row=summary_row + 3, column=1, value="Total FTE:")
    total_fte = sum(row[4] for row in portfolio_data)
    ws.cell(row=summary_row + 3, column=2, value=total_fte)
    
    # Unique headquarters
    unique_hq = len(set(row[5] for row in portfolio_data))
    ws.cell(row=summary_row + 4, column=1, value="Unique Headquarters:")
    ws.cell(row=summary_row + 4, column=2, value=unique_hq)
    
    # Industry breakdown
    industries = {}
    for row in portfolio_data:
        industry = row[1]
        industries[industry] = industries.get(industry, 0) + 1
    
    ws.cell(row=summary_row + 6, column=1, value="Industry Breakdown:").font = Font(bold=True)
    breakdown_row = summary_row + 7
    for industry, count in sorted(industries.items()):
        ws.cell(row=breakdown_row, column=1, value=f"  {industry}:")
        ws.cell(row=breakdown_row, column=2, value=count)
        breakdown_row += 1
    
    # Save the workbook
    output_path = "./Assets/Portfolio/Portfolio_Overview.xlsx"
    wb.save(output_path)
    print(f"Excel file created successfully at: {output_path}")
    
    return output_path

if __name__ == "__main__":
    create_portfolio_excel()
